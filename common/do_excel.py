
import openpyxl


class Cases:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.method = None
        self.url = None
        self.date = None
        self.expected = None


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.wb = openpyxl.load_workbook(self.file_name)
        self.sheet = self.wb[self.sheet_name]

    def raed(self):
        cases = []
        for i in range(2, self.sheet.max_row+1):
            case_date = Cases()
            case_date.case_id = self.sheet.cell(row=i, column=1).value
            case_date.title = self.sheet.cell(row=i, column=2).value
            case_date.method = self.sheet.cell(row=i, column=3).value
            case_date.url = self.sheet.cell(row=i, column=4).value
            case_date.date = self.sheet.cell(row=i, column=5).value
            case_date.expected = self.sheet.cell(row=i, column=6).value
            cases.append(case_date)
        return cases

    def write(self, row, actual, result):
        self.sheet.cell(row=row, column=7).value = actual
        self.sheet.cell(row=row, column=8).value = result
        self.wb.save(self.file_name)

if __name__ == '__main__':
    from common import contants
    do_excel = DoExcel(contants.data_path, 'register')
    cases = do_excel.raed()
    print(cases[1].date)
